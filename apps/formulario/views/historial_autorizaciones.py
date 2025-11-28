from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from django.views import View
from django.http import HttpResponse
from apps.formulario.models import HistorialAutorizacion, TipoAutorizacion
from django.utils import timezone
from django.db.models import Q
from django.conf import settings
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import openpyxl
import os

# ============================================================================
# HISTORIAL DE AUTORIZACIONES
# ============================================================================

class HistorialAutorizacionListView(LoginRequiredMixin, ListView):
    """Lista de historial de autorizaciones con filtros para reportes"""
    model = HistorialAutorizacion
    template_name = 'formulario/historial_autorizaciones_list.html'
    context_object_name = 'historial_autorizaciones'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = HistorialAutorizacion.objects.select_related(
            'autorizacion',
            'autorizacion__usuario',
            'autorizacion__tipo_autorizacion',
            'creado_por'
        )
        
        # Filtro por RANGO de fecha de creación (emisión)
        fecha_creacion_desde = self.request.GET.get('fecha_creacion_desde')
        fecha_creacion_hasta = self.request.GET.get('fecha_creacion_hasta')
        
        if fecha_creacion_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_creacion_desde, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_creacion__date__gte=fecha_desde_obj)
            except ValueError:
                pass
        
        if fecha_creacion_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_creacion_hasta, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_creacion__date__lte=fecha_hasta_obj)
            except ValueError:
                pass
        
        # Filtro por RANGO de fecha de vigencia (caducidad)
        fecha_vigencia_desde = self.request.GET.get('fecha_vigencia_desde')
        fecha_vigencia_hasta = self.request.GET.get('fecha_vigencia_hasta')
        
        if fecha_vigencia_desde:
            try:
                fecha_vigencia_desde_obj = datetime.strptime(fecha_vigencia_desde, '%Y-%m-%d').date()
                queryset = queryset.filter(autorizacion__vigencia__gte=fecha_vigencia_desde_obj)
            except ValueError:
                pass
        
        if fecha_vigencia_hasta:
            try:
                fecha_vigencia_hasta_obj = datetime.strptime(fecha_vigencia_hasta, '%Y-%m-%d').date()
                queryset = queryset.filter(autorizacion__vigencia__lte=fecha_vigencia_hasta_obj)
            except ValueError:
                pass
        
        # Filtro por tipo de autorización
        tipo_autorizacion = self.request.GET.get('tipo_autorizacion')
        if tipo_autorizacion:
            queryset = queryset.filter(autorizacion__tipo_autorizacion_id=tipo_autorizacion)
        
        # Filtro por placa
        placa = self.request.GET.get('placa')
        if placa:
            queryset = queryset.filter(autorizacion__placa__icontains=placa)
        
        # Filtro por usuario
        usuario = self.request.GET.get('usuario')
        if usuario:
            queryset = queryset.filter(autorizacion__usuario__nombres__icontains=usuario)
        
        # Filtro por estado
        estado = self.request.GET.get('estado')
        if estado == 'vigentes':
            queryset = queryset.filter(autorizacion__vigencia__gte=timezone.now().date())
        elif estado == 'caducadas':
            queryset = queryset.filter(autorizacion__vigencia__lt=timezone.now().date())
        
        return queryset.order_by('-fecha_creacion')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Tipos de autorización para el filtro
        context['tipos_autorizacion'] = TipoAutorizacion.objects.filter(activo=True)
        
        # Total de registros
        context['total_registros'] = self.get_queryset().count()
        
        # Estadísticas
        queryset = self.get_queryset()
        context['total_vigentes'] = queryset.filter(
            autorizacion__vigencia__gte=timezone.now().date()
        ).count()
        context['total_caducadas'] = queryset.filter(
            autorizacion__vigencia__lt=timezone.now().date()
        ).count()
        
        # Preservar valores de filtros en el contexto
        context['filtros'] = {
            'fecha_creacion_desde': self.request.GET.get('fecha_creacion_desde', ''),
            'fecha_creacion_hasta': self.request.GET.get('fecha_creacion_hasta', ''),
            'fecha_vigencia_desde': self.request.GET.get('fecha_vigencia_desde', ''),
            'fecha_vigencia_hasta': self.request.GET.get('fecha_vigencia_hasta', ''),
            'tipo_autorizacion': self.request.GET.get('tipo_autorizacion', ''),
            'placa': self.request.GET.get('placa', ''),
            'usuario': self.request.GET.get('usuario', ''),
            'estado': self.request.GET.get('estado', ''),
        }
        
        context['current_date'] = timezone.now().strftime('%d/%m/%Y, %H:%M')
        return context


class ExportarHistorialExcelView(LoginRequiredMixin, View):
    """Exportar historial de autorizaciones a Excel"""
    
    def get(self, request, *args, **kwargs):
        # Obtener el queryset con los mismos filtros de la lista
        queryset = HistorialAutorizacion.objects.select_related(
            'autorizacion',
            'autorizacion__usuario',
            'autorizacion__tipo_autorizacion',
            'creado_por'
        )
        
        # Aplicar filtros (igual que en ListView) - ACTUALIZADO CON RANGOS
        
        # Filtro por RANGO de fecha de creación (emisión)
        fecha_creacion_desde = request.GET.get('fecha_creacion_desde')
        fecha_creacion_hasta = request.GET.get('fecha_creacion_hasta')
        
        if fecha_creacion_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_creacion_desde, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_creacion__date__gte=fecha_desde_obj)
            except ValueError:
                pass
        
        if fecha_creacion_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_creacion_hasta, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_creacion__date__lte=fecha_hasta_obj)
            except ValueError:
                pass
        
        # Filtro por RANGO de fecha de vigencia (caducidad)
        fecha_vigencia_desde = request.GET.get('fecha_vigencia_desde')
        fecha_vigencia_hasta = request.GET.get('fecha_vigencia_hasta')
        
        if fecha_vigencia_desde:
            try:
                fecha_vigencia_desde_obj = datetime.strptime(fecha_vigencia_desde, '%Y-%m-%d').date()
                queryset = queryset.filter(autorizacion__vigencia__gte=fecha_vigencia_desde_obj)
            except ValueError:
                pass
        
        if fecha_vigencia_hasta:
            try:
                fecha_vigencia_hasta_obj = datetime.strptime(fecha_vigencia_hasta, '%Y-%m-%d').date()
                queryset = queryset.filter(autorizacion__vigencia__lte=fecha_vigencia_hasta_obj)
            except ValueError:
                pass
        
        tipo_autorizacion = request.GET.get('tipo_autorizacion')
        if tipo_autorizacion:
            queryset = queryset.filter(autorizacion__tipo_autorizacion_id=tipo_autorizacion)
        
        placa = request.GET.get('placa')
        if placa:
            queryset = queryset.filter(autorizacion__placa__icontains=placa)
        
        usuario = request.GET.get('usuario')
        if usuario:
            queryset = queryset.filter(autorizacion__usuario__nombres__icontains=usuario)
        
        estado = request.GET.get('estado')
        if estado == 'vigentes':
            queryset = queryset.filter(autorizacion__vigencia__gte=timezone.now().date())
        elif estado == 'caducadas':
            queryset = queryset.filter(autorizacion__vigencia__lt=timezone.now().date())
        
        queryset = queryset.order_by('-fecha_creacion')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Autorizaciones"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="004d99", end_color="004d99", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        title_font = Font(bold=True, size=16, color="004d99")
        subtitle_font = Font(size=11, color="666666")
        info_font = Font(bold=True, size=10)
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ========== AGREGAR LOGO ==========
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'Logo_ACT.png')
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 100
            img.height = 100
            ws.add_image(img, 'A1')
        
        # Ajustar altura de las primeras filas para el logo
        ws.row_dimensions[1].height = 50
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 20
        
        # ========== ENCABEZADO DEL REPORTE ==========
        ws.merge_cells('C1:H1')
        ws['C1'] = 'REPORTE DE HISTORIAL DE AUTORIZACIONES'
        ws['C1'].font = title_font
        ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells('C2:H2')
        ws['C2'] = 'EMOVIM-EP | Autoridad de Control de Tránsito Milagro'
        ws['C2'].font = subtitle_font
        ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Fecha de generación
        ws.merge_cells('A3:H3')
        ws['A3'] = f'Generado el: {timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M:%S")}'
        ws['A3'].font = subtitle_font
        ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Mostrar rangos de fechas aplicados
        filtros_texto = []
        if fecha_creacion_desde or fecha_creacion_hasta:
            rango_creacion = "Fecha Emisión: "
            if fecha_creacion_desde:
                rango_creacion += f"Desde {datetime.strptime(fecha_creacion_desde, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            if fecha_creacion_hasta:
                if fecha_creacion_desde:
                    rango_creacion += f" hasta {datetime.strptime(fecha_creacion_hasta, '%Y-%m-%d').strftime('%d/%m/%Y')}"
                else:
                    rango_creacion += f"Hasta {datetime.strptime(fecha_creacion_hasta, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            filtros_texto.append(rango_creacion)
        
        if fecha_vigencia_desde or fecha_vigencia_hasta:
            rango_vigencia = "Vigencia: "
            if fecha_vigencia_desde:
                rango_vigencia += f"Desde {datetime.strptime(fecha_vigencia_desde, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            if fecha_vigencia_hasta:
                if fecha_vigencia_desde:
                    rango_vigencia += f" hasta {datetime.strptime(fecha_vigencia_hasta, '%Y-%m-%d').strftime('%d/%m/%Y')}"
                else:
                    rango_vigencia += f"Hasta {datetime.strptime(fecha_vigencia_hasta, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            filtros_texto.append(rango_vigencia)
        
        # Mostrar filtros aplicados
        if filtros_texto:
            ws.merge_cells('A4:H4')
            ws['A4'] = f'Filtros aplicados: {" | ".join(filtros_texto)}'
            ws['A4'].font = Font(size=9, italic=True, color="666666")
            ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Total de registros
        ws.merge_cells('A5:H5')
        ws['A5'] = f'Total de registros: {queryset.count()}'
        ws['A5'].font = info_font
        ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Fondo azul claro para el área de encabezado
        light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        for row in range(1, 6):
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                if row > 1 or col > 2:
                    cell.fill = light_blue_fill
        
        # Encabezados de columnas (ahora en la fila 7)
        headers = [
            'Fecha de Emisión',
            'Hora',
            'Fecha de Vigencia',
            'Tipo de Autorización',
            'Placa',
            'Usuario Autorizado',
            'Número de Autorización',
            'Estado'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin
        
        ws.row_dimensions[7].height = 30
        
        # Datos (ahora empiezan en la fila 8)
        row_num = 8
        for historial in queryset:
            ws.cell(row=row_num, column=1, value=historial.fecha_creacion.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=2, value=historial.fecha_creacion.strftime('%H:%M:%S'))
            ws.cell(row=row_num, column=3, value=historial.autorizacion.vigencia.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=4, value=historial.autorizacion.tipo_autorizacion.nombre)
            ws.cell(row=row_num, column=5, value=historial.autorizacion.placa)
            ws.cell(row=row_num, column=6, value=historial.autorizacion.usuario.nombres)
            ws.cell(row=row_num, column=7, value=historial.autorizacion.numero_autorizacion)
            
            # Estado
            estado_valor = 'VIGENTE' if not historial.autorizacion.esta_caducada else 'CADUCADA'
            cell_estado = ws.cell(row=row_num, column=8, value=estado_valor)
            
            # Colorear según estado
            if historial.autorizacion.esta_caducada:
                cell_estado.font = Font(color="FF0000", bold=True)
            else:
                cell_estado.font = Font(color="008000", bold=True)
            
            # Aplicar bordes y alineación
            for col in range(1, 9):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Aplicar fondo alternado (estilo zebra)
            if row_num % 2 == 0:
                zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = zebra_fill
            
            row_num += 1
        
        # ========== PIE DE PÁGINA ==========
        footer_row = row_num + 1
        ws.merge_cells(f'A{footer_row}:H{footer_row}')
        ws[f'A{footer_row}'] = '© EMOVIM-EP | Av. Simón Bolívar y Juan Montalvo | Milagro, Ecuador | genercia@emovim-ep.gob.ec'
        ws[f'A{footer_row}'].font = Font(size=9, color="666666", italic=True)
        ws[f'A{footer_row}'].alignment = Alignment(horizontal="center", vertical="center")
        
        footer_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws[f'A{footer_row}'].fill = footer_fill
        ws.row_dimensions[footer_row].height = 25
        
        # Ajustar ancho de columnas
        column_widths = [18, 12, 18, 25, 15, 30, 30, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_hora_local = timezone.localtime(timezone.now())
        filename = f'Historial_Autorizaciones_{fecha_hora_local.strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response